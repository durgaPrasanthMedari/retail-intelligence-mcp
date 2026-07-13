from mcp.server.fastmcp import FastMCP
import sqlite3
import json
import os

# ── Initialise server ─────────────────────────────────
mcp = FastMCP("Retail Intelligence")
DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "retail.db")

# ── Helper: connect and return dict rows ──────────────
def query(sql: str, params: tuple = ()):
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def execute(sql: str, params: tuple = ()):
    """Run INSERT/UPDATE/CREATE — no return value needed."""
    conn = sqlite3.connect(DB)
    conn.execute(sql, params)
    conn.commit()
    conn.close()

# ══ TOOL 1: Get all products ══════════════════════════
@mcp.tool()
def get_all_products() -> str:
    """Get the complete product list with SKU, name,
    category, price, and current stock levels."""
    results = query("SELECT sku, name, category, price, stock FROM products ORDER BY category")
    return json.dumps(results, indent=2)

# ══ TOOL 2: Sell-through report ═══════════════════════
@mcp.tool()
def get_sell_through_report(category: str = "all") -> str:
    """Get sell-through percentage for all products or
    filter by category. Categories: Outerwear, Knitwear,
    Trousers, Shoes, Accessories. Use 'all' for everything.
    Returns sku, name, sell_through_pct, and stock remaining."""
    if category == "all":
        results = query("""
            SELECT sku, name, category,
                   ROUND(units_sold*100.0/total_units,1) AS sell_through_pct,
                   stock AS remaining_stock
            FROM products ORDER BY sell_through_pct DESC
        """)
    else:
        results = query("""
            SELECT sku, name, category,
                   ROUND(units_sold*100.0/total_units,1) AS sell_through_pct,
                   stock AS remaining_stock
            FROM products WHERE category = ?
            ORDER BY sell_through_pct DESC
        """, (category,))
    return json.dumps(results, indent=2)

# ══ TOOL 3: Find underperformers ══════════════════════
@mcp.tool()
def find_underperformers(threshold_pct: float = 50.0) -> str:
    """Find products with sell-through below a threshold.
    Default threshold is 50%. Returns SKU, name, category,
    sell_through_pct, and stock_value_at_risk in euros.
    Use lower threshold (e.g. 30) to find worst performers."""
    results = query("""
        SELECT sku, name, category,
               ROUND(units_sold*100.0/total_units,1) AS sell_through_pct,
               stock AS units_remaining,
               ROUND(price*stock,2) AS stock_value_at_risk_eur
        FROM products
        WHERE (units_sold*100.0/total_units) < ?
        ORDER BY sell_through_pct ASC
    """, (threshold_pct,))
    if not results:
        return f"No products found below {threshold_pct}% sell-through."
    return json.dumps(results, indent=2)

# ══ TOOL 4: Get suppliers ═════════════════════════════
@mcp.tool()
def get_suppliers() -> str:
    """Get all suppliers with their country, lead time
    in days, and reliability rating."""
    results = query("SELECT * FROM suppliers ORDER BY lead_days")
    return json.dumps(results, indent=2)

# ══ TOOL 5: Weekly sales summary ══════════════════════
@mcp.tool()
def get_weekly_sales(week: str = "W48") -> str:
    """Get sales performance for a specific week.
    Week format: W48, W49 etc. Returns product name,
    units sold, revenue, and average selling price."""
    results = query("""
        SELECT ws.sku, p.name, p.category,
               ws.units_sold, ws.revenue,
               ROUND(ws.revenue/ws.units_sold,2) AS avg_price
        FROM weekly_sales ws
        JOIN products p ON ws.sku = p.sku
        WHERE ws.week = ?
        ORDER BY ws.revenue DESC
    """, (week,))
    if not results:
        return f"No sales data found for week {week}."
    total_rev = sum(r['revenue'] for r in results)
    total_units = sum(r['units_sold'] for r in results)
    return json.dumps({
        "week": week,
        "total_revenue_eur": round(total_rev, 2),
        "total_units_sold": total_units,
        "products": results
    }, indent=2)

# ══ TOOL 6: Low stock alert ════════════════════════════
@mcp.tool()
def get_low_stock_alerts(weeks_threshold: int = 4) -> str:
    """Find products that are close to selling out.
    Returns products where estimated weeks of stock
    remaining is below the threshold. Default is 4 weeks.
    Based on average weekly sales rate."""
    results = query("""
        SELECT p.sku, p.name, p.category, p.stock,
               ROUND(ws.avg_weekly, 1) AS avg_weekly_sales,
               ROUND(p.stock / NULLIF(ws.avg_weekly, 0), 1) AS weeks_of_stock
        FROM products p
        LEFT JOIN (
            SELECT sku, AVG(units_sold) AS avg_weekly
            FROM weekly_sales GROUP BY sku
        ) ws ON p.sku = ws.sku
        WHERE (p.stock / NULLIF(ws.avg_weekly, 0)) < ?
        OR ws.avg_weekly IS NULL
        ORDER BY weeks_of_stock ASC
    """, (weeks_threshold,))
    return json.dumps(results, indent=2)

# ══ TOOL 7: Rate of sale with weeks cover ════════════════
@mcp.tool()
def get_rate_of_sale(category: str = "all") -> str:
    """Calculate weekly rate of sale and weeks of stock cover
    for all products or filtered by category. Returns avg_weekly_sales,
    weeks_cover, sell_through_pct, and stock_value_eur.
    Critical for reorder and markdown decisions."""
    sql = """
        SELECT
            p.sku, p.name, p.category,
            p.price, p.stock,
            ROUND(p.units_sold * 100.0 / p.total_units, 1) AS sell_through_pct,
            ROUND(COALESCE(ws.avg_weekly, 0), 1) AS avg_weekly_sales,
            CASE
                WHEN COALESCE(ws.avg_weekly, 0) = 0 THEN 999
                ELSE ROUND(p.stock / ws.avg_weekly, 1)
            END AS weeks_cover,
            ROUND(p.price * p.stock, 2) AS stock_value_eur,
            CASE
                WHEN (p.units_sold*100.0/p.total_units) >= 80 THEN 'Hero'
                WHEN (p.units_sold*100.0/p.total_units) >= 50 THEN 'On track'
                WHEN (p.units_sold*100.0/p.total_units) >= 30 THEN 'Watch'
                ELSE 'Action needed'
            END AS status
        FROM products p
        LEFT JOIN (
            SELECT sku, AVG(units_sold) AS avg_weekly
            FROM weekly_sales GROUP BY sku
        ) ws ON p.sku = ws.sku
        WHERE (p.category = ? OR ? = 'all')
        ORDER BY weeks_cover ASC
    """
    results = query(sql, (category, category))
    total_value = sum(r['stock_value_eur'] for r in results)
    return json.dumps({
        "summary": {
            "total_skus": len(results),
            "total_stock_value_eur": round(total_value, 2),
            "heroes": sum(1 for r in results if r['status'] == 'Hero'),
            "action_needed": sum(1 for r in results if r['status'] == 'Action needed')
        },
        "products": results
    }, indent=2)

# ══ TOOL 8: Markdown candidates with urgency score ═══════
@mcp.tool()
def get_markdown_candidates(urgency_threshold: int = 50) -> str:
    """Score every SKU by markdown urgency (0-100 scale).
    Higher score = more urgent to markdown. Score is based on
    sell-through rate, weeks of cover, and stock value at risk.
    Use urgency_threshold=50 for moderately urgent, 70 for critical."""
    results = query("""
        SELECT p.sku, p.name, p.category,
               ROUND(p.units_sold*100.0/p.total_units,1) AS sell_through_pct,
               p.stock, p.price,
               ROUND(p.price*p.stock,2) AS stock_value_eur,
               ROUND(COALESCE(ws.avg_weekly,0),1) AS avg_weekly_sales,
               CASE WHEN COALESCE(ws.avg_weekly,0)=0 THEN 999
                    ELSE ROUND(p.stock/ws.avg_weekly,1) END AS weeks_cover
        FROM products p
        LEFT JOIN (
            SELECT sku, AVG(units_sold) AS avg_weekly
            FROM weekly_sales GROUP BY sku
        ) ws ON p.sku = ws.sku
    """)
    scored = []
    for r in results:
        st_score = max(0, 60 - r['sell_through_pct'])
        wc = r['weeks_cover']
        wc_score = max(0, min(30, (20 - wc) * 1.5)) if wc < 999 else 0
        val_score = min(10, r['stock_value_eur'] / 2000)
        urgency = round(min(100, st_score + wc_score + val_score))
        if urgency >= urgency_threshold:
            r['urgency_score'] = urgency
            r['recommendation'] = (
                "Immediate 30%+ markdown" if urgency >= 75
                else "Planned 15-20% markdown" if urgency >= 50
                else "Monitor — markdown ready"
            )
            scored.append(r)
    scored.sort(key=lambda x: x['urgency_score'], reverse=True)
    return json.dumps(scored, indent=2)

# ══ TOOL 9: Category performance ═════════════════════════
@mcp.tool()
def get_category_performance() -> str:
    """Compare performance across all product categories.
    Returns avg sell-through, total revenue, stock value,
    best and worst performing SKU per category."""
    results = query("""
        SELECT
            category,
            COUNT(*) AS sku_count,
            ROUND(AVG(units_sold*100.0/total_units),1) AS avg_sell_through_pct,
            ROUND(SUM(units_sold*price),2) AS total_revenue_eur,
            ROUND(SUM(stock*price),2) AS stock_value_at_risk_eur,
            MIN(ROUND(units_sold*100.0/total_units,1)) AS worst_st_pct,
            MAX(ROUND(units_sold*100.0/total_units,1)) AS best_st_pct
        FROM products
        GROUP BY category
        ORDER BY avg_sell_through_pct DESC
    """)
    return json.dumps(results, indent=2)

# ══ TOOL 10: Generate Excel report ═══════════════════════
@mcp.tool()
def generate_excel_report(filename: str = "retail_report.xlsx") -> str:
    """Generate a formatted Excel report with 3 sheets:
    Summary dashboard, full sell-through analysis, and
    markdown candidates. Saves the file and returns the path.
    Client can open it directly in Excel."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    GREEN = PatternFill("solid", fgColor="C6EFCE")
    AMBER = PatternFill("solid", fgColor="FFEB9C")
    RED   = PatternFill("solid", fgColor="FFC7CE")
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header_row(ws, row, cols):
        for col, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
            c.alignment = center; c.border = thin

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws1 = wb.active; ws1.title = "Summary"
    ws1["A1"] = "Retail Intelligence Report"
    ws1["A1"].font = Font(bold=True, size=16)
    ws1["A2"] = f"Generated: {__import__('datetime').datetime.now().strftime('%d %b %Y %H:%M')}"
    cats = query("""SELECT category,
        ROUND(AVG(units_sold*100.0/total_units),1) AS avg_st,
        ROUND(SUM(stock*price),2) AS stock_val
        FROM products GROUP BY category ORDER BY avg_st DESC""")
    style_header_row(ws1, 4, ["Category","Avg Sell-Through %","Stock Value €"])
    for i, r in enumerate(cats, 5):
        ws1.cell(i,1,r['category']).border=thin
        c2=ws1.cell(i,2,r['avg_st']); c2.border=thin; c2.alignment=center
        c2.fill = GREEN if r['avg_st']>=70 else AMBER if r['avg_st']>=45 else RED
        c3=ws1.cell(i,3,r['stock_val']); c3.border=thin; c3.number_format='#,##0.00'
    ws1.column_dimensions['A'].width=16
    ws1.column_dimensions['B'].width=22
    ws1.column_dimensions['C'].width=16

    # Sheet 2: Sell-Through
    ws2 = wb.create_sheet("Sell-Through")
    style_header_row(ws2, 1, ["SKU","Name","Category","Price €","Total Units","Sold","Stock","ST %","Stock Value €"])
    products = query("SELECT sku,name,category,price,total_units,units_sold,stock FROM products ORDER BY category,units_sold*100.0/total_units DESC")
    for i, p in enumerate(products, 2):
        st = round(p['units_sold']*100.0/p['total_units'],1)
        vals=[p['sku'],p['name'],p['category'],p['price'],p['total_units'],p['units_sold'],p['stock'],st,round(p['price']*p['stock'],2)]
        for j,v in enumerate(vals,1):
            c=ws2.cell(i,j,v); c.border=thin
        st_cell=ws2.cell(i,8); st_cell.alignment=center
        st_cell.fill = GREEN if st>=70 else AMBER if st>=45 else RED
    for col,w in zip("ABCDEFGHI",[8,22,12,9,12,8,8,8,14]):
        ws2.column_dimensions[col].width=w

    # Sheet 3: Markdown Candidates
    ws3 = wb.create_sheet("Markdown Candidates")
    style_header_row(ws3, 1, ["SKU","Name","Category","ST %","Stock","Stock Value €","Action"])
    underperformers = query("""SELECT sku,name,category,
        ROUND(units_sold*100.0/total_units,1) AS st,
        stock, ROUND(price*stock,2) AS val
        FROM products WHERE (units_sold*100.0/total_units)<50
        ORDER BY units_sold*100.0/total_units ASC""")
    for i, p in enumerate(underperformers, 2):
        action = "Immediate markdown 30%+" if p['st']<30 else "Planned markdown 15-20%"
        vals=[p['sku'],p['name'],p['category'],p['st'],p['stock'],p['val'],action]
        for j,v in enumerate(vals,1):
            c=ws3.cell(i,j,v); c.border=thin; c.fill=RED

    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb.save(path)
    return f"Excel report saved to: {path}\nSheets: Summary, Sell-Through, Markdown Candidates\nOpen it directly in Excel."

# ══ TOOL 11: Weather forecast + demand prediction ════════
# FIX: Removed broken import of query_db.get_top_selling_cities().
# Kept your multi-city concept — now works by accepting a comma-separated
# list of cities, defaulting to Milan if none provided.
@mcp.tool()
def get_weather_and_demand_forecast(
    cities: str = "Milan",
    days: int = 7
) -> str:
    """Get weather forecast and predict demand impact on fashion categories.
    Pass one city or multiple comma-separated cities e.g. 'Milan,Rome,Florence'.
    Defaults to Milan if omitted. Uses Open-Meteo free API — no key needed.
    Returns daily forecast with temperature, rain, and strategic recommendations."""
    import requests

    target_cities = [c.strip() for c in cities.split(",") if c.strip()]
    if not target_cities:
        target_cities = ["Milan"]

    def demand_signals(tmax, tmin, rain_mm, wind_kmh):
        avg = (tmax + tmin) / 2
        return {
            "Outerwear": (
                "HIGH — cold snap or rain" if avg < 8 or rain_mm > 10 or wind_kmh > 40
                else "MODERATE" if avg < 15
                else "LOW — too warm"
            ),
            "Knitwear": (
                "HIGH — ideal knitwear weather" if 6 < avg < 18
                else "MODERATE" if avg < 22
                else "LOW — too warm"
            ),
            "Shoes": "Push boots — rain expected" if rain_mm > 5 else "Normal",
            "Accessories": "HIGH — scarves and belts" if avg < 12 else "Normal",
            "Trousers": "NORMAL — not weather-sensitive"
        }

    final_report = {
        "analysis_type": "Multi-City" if len(target_cities) > 1 else "Single City",
        "results": []
    }

    for current_city in target_cities:
        try:
            geo = requests.get(
                "https://geocoding-api.open-meteo.com/v1/search",
                params={"name": current_city, "count": 1},
                timeout=10
            ).json()
        except Exception as e:
            final_report["results"].append({"city": current_city, "error": str(e)})
            continue

        if not geo.get("results"):
            final_report["results"].append({"city": current_city, "error": "City not found"})
            continue

        lat = geo["results"][0]["latitude"]
        lon = geo["results"][0]["longitude"]
        tz  = geo["results"][0].get("timezone", "Europe/Rome")

        try:
            resp = requests.get(
                "https://api.open-meteo.com/v1/forecast",
                params={
                    "latitude": lat, "longitude": lon,
                    "daily": ["temperature_2m_max","temperature_2m_min","precipitation_sum","windspeed_10m_max"],
                    "forecast_days": days,
                    "timezone": tz
                },
                timeout=10
            ).json()
        except Exception as e:
            final_report["results"].append({"city": current_city, "error": f"Weather API: {str(e)}"})
            continue

        daily = resp.get("daily", {})
        if not daily:
            continue

        dates = daily["time"]
        t_max = daily["temperature_2m_max"]
        t_min = daily["temperature_2m_min"]
        rain  = daily["precipitation_sum"]
        wind  = daily["windspeed_10m_max"]

        result_days = []
        for i in range(len(dates)):
            avg_t = round((t_max[i] + t_min[i]) / 2, 1)
            result_days.append({
                "date": dates[i],
                "avg_temp_c": avg_t,
                "rain_mm": rain[i],
                "demand_signals": demand_signals(t_max[i], t_min[i], rain[i], wind[i])
            })

        avg_overall_temp = sum(t_max) / len(t_max)
        total_rain = sum(rain)
        strategies = []
        if avg_overall_temp > 22:
            strategies.append("Markdown Alert: Promote heavy Outerwear/Knitwear — too warm for full price.")
        if total_rain > 12:
            strategies.append("Logistics Action: Move waterproof footwear to high-visibility zones.")
        if 8 <= avg_overall_temp <= 16:
            strategies.append("Margin Protection: Keep full pricing on Knitwear — ideal weather window.")

        final_report["results"].append({
            "city": current_city,
            "weather_summary": {
                "avg_max_temp_c": round(avg_overall_temp, 1),
                "total_rain_mm": round(total_rain, 1)
            },
            "strategic_recommendations": strategies,
            "daily_forecast": result_days
        })

    return json.dumps(final_report, indent=2)

# ══ TOOL 12: Capture and analyse a new trend ═════════════
# FIX: Removed broken import of query_db.save_new_market_trend().
# Kept your excellent concept — now creates the trends table
# automatically and uses the existing query() helper.
@mcp.tool()
def capture_and_analyze_trend(
    trend_name: str,
    description: str,
    source: str = "Social Media / TikTok",
    relevant_keywords: list = None
) -> str:
    """Ingest a newly discovered fashion trend or viral event.
    Logs it to the trends table in the database and immediately
    cross-references inventory to find matching stock.
    Use for: TikTok viral moments, runway trends, seasonal shifts."""
    from datetime import datetime

    # Auto-create trends table if it doesn't exist yet
    execute("""
        CREATE TABLE IF NOT EXISTS trends (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trend_name TEXT NOT NULL,
            source TEXT,
            description TEXT,
            logged_at TEXT,
            status TEXT DEFAULT 'active'
        )
    """)

    # Save to database
    execute(
        "INSERT INTO trends (trend_name, source, description, logged_at) VALUES (?,?,?,?)",
        (trend_name, source, description, datetime.now().strftime("%Y-%m-%d %H:%M"))
    )

    # Match against inventory using provided keywords or words from trend name
    keywords = relevant_keywords if relevant_keywords else trend_name.split()
    matching_stock = []
    for kw in keywords:
        matches = query("""
            SELECT sku, name, category, stock, price,
                   ROUND(units_sold*100.0/total_units,1) AS sell_through_pct
            FROM products
            WHERE LOWER(name) LIKE ? OR LOWER(category) LIKE ?
        """, (f"%{kw.lower()}%", f"%{kw.lower()}%"))
        for m in matches:
            if m not in matching_stock:
                matching_stock.append(m)

    return json.dumps({
        "trend_registered": trend_name,
        "source": source,
        "description": description,
        "matched_inventory_count": len(matching_stock),
        "inventory_matches": matching_stock,
        "strategic_recommendation": (
            "High inventory match — create an immediate marketing push."
            if len(matching_stock) > 0
            else "No stock match. Consider capsule buy or vendor sourcing for next season."
        )
    }, indent=2)

# ══ TOOL 13: Draft trend alert email ═════════════════════
# YOUR TOOL — works correctly, kept as-is.
# Drafts the email content but does not send. Use send_alert_email to send.
@mcp.tool()
def draft_trend_alert_email(
    trend_name: str,
    recipient_team: str = "Marketing & E-Commerce",
    action_plan_summary: str = ""
) -> str:
    """Draft an urgent trend alert email for regional teams, buyers,
    or store managers. Returns a ready-to-send email draft with
    subject and body. Does not send — use send_alert_email to dispatch."""
    from datetime import datetime
    today = datetime.now().strftime("%B %d, %Y")
    subject = f"INTERNAL FLASH BRIEFING: Capitalizing on the '{trend_name}' Movement"
    body = (
        f"To: {recipient_team} Team\n"
        f"From: Merchandising Analytics & Strategy AI\n"
        f"Date: {today}\n\n"
        f"Dear Team,\n\n"
        f"We have identified a significant market momentum shift: '{trend_name}'.\n"
        f"To maximise gross margin and capture early consumer demand, act immediately.\n\n"
        f"PROPOSED ACTIONS:\n"
        f"{action_plan_summary if action_plan_summary else '- Pivot homepage carousels to showcase related items.\n- Rearrange front-of-store racks to match the trend aesthetic.\n- Brief social media team on content angle.'}\n\n"
        f"Please review stock metrics and adjust promotional activity accordingly.\n\n"
        f"Regards,\nMerchandising Planning — AI Intelligence Layer"
    )
    return json.dumps({
        "subject": subject,
        "recipient": recipient_team,
        "generated_draft": body
    }, indent=2)

# ══ TOOL 14: Scan fashion trends via DuckDuckGo ══════════
# YOUR VERSION — cleaner than the original. Kept as-is.
@mcp.tool()
def scan_fashion_trends(
    search_term: str = "fashion trends autumn winter 2026",
    max_results: int = 6
) -> str:
    """Search the web for current fashion trends using DuckDuckGo.
    No API key needed. Returns article titles, summaries and URLs.
    Use to find trending styles, colours, or viral fashion moments.
    Then use analyse_trend_vs_stock to cross-reference your inventory."""
    from duckduckgo_search import DDGS

    try:
        results = []
        ddgs = DDGS()
        search_results = ddgs.text(keywords=search_term, max_results=max_results)
        for r in search_results:
            results.append({
                "title":   r.get("title", "No Title"),
                "summary": r.get("body", "")[:400],
                "url":     r.get("href", "")
            })
        return json.dumps({
            "search_term":   search_term,
            "results_found": len(results),
            "articles":      results,
            "tip": "Use analyse_trend_vs_stock with a keyword from these results"
        }, indent=2)
    except Exception as e:
        return json.dumps({
            "error": f"Search failed: {str(e)}",
            "tip": "Try again in 10 seconds — DuckDuckGo occasionally rate-limits."
        }, indent=2)

# ══ TOOL 15: Analyse trend vs stock ══════════════════════
@mcp.tool()
def analyse_trend_vs_stock(trend_keyword: str) -> str:
    """Check which products in current stock align with a trend keyword.
    Searches product names and categories. Returns matching SKUs with
    stock levels and sell-through, plus gap analysis for buying."""
    kw = f"%{trend_keyword.lower()}%"
    matches = query("""
        SELECT sku, name, category, stock, price,
               ROUND(units_sold * 100.0 / total_units, 1) AS sell_through_pct,
               ROUND(price * stock, 2) AS stock_value_eur
        FROM products
        WHERE LOWER(name) LIKE ? OR LOWER(category) LIKE ?
        ORDER BY sell_through_pct DESC
    """, (kw, kw))

    high_stock = [p for p in matches if p['stock'] > 30]
    low_st     = [p for p in matches if p['sell_through_pct'] < 50]

    return json.dumps({
        "trend_keyword": trend_keyword,
        "matching_skus": len(matches),
        "products": matches,
        "opportunity": {
            "push_in_marketing":  [p['name'] for p in high_stock],
            "slow_movers_to_tie": [p['name'] for p in low_st],
            "buying_gap": (
                "No current stock — consider for next season buy."
                if not matches
                else "Stock available — leverage this trend in marketing now."
            )
        }
    }, indent=2)

# ══ TOOL 16: Send alert email ═════════════════════════════
@mcp.tool()
def send_alert_email(
    to_email: str,
    subject: str,
    body_text: str
) -> str:
    """Send a branded HTML email alert to any address.
    Use for: weekly trading briefs, markdown urgency alerts,
    low-stock warnings, and trend notifications.
    Reads Gmail credentials from the .env file automatically."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from dotenv import load_dotenv

    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    load_dotenv(env_path)

    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_APP_PASSWORD")

    if not gmail_user or not gmail_pass:
        return "Email not configured. Add GMAIL_USER and GMAIL_APP_PASSWORD to your .env file."

    html = f"""<html>
<body style="font-family:Arial,sans-serif;max-width:600px;margin:auto">
  <div style="background:#1F4E79;padding:20px 24px;border-radius:8px 8px 0 0">
    <h2 style="color:white;margin:0;font-size:18px">Retail Intelligence</h2>
    <p style="color:#a8c8e8;margin:4px 0 0;font-size:13px">{subject}</p>
  </div>
  <div style="padding:24px;background:#ffffff;border:1px solid #dde">
    <pre style="font-family:Arial;font-size:14px;white-space:pre-wrap;
                line-height:1.6;color:#333">{body_text}</pre>
  </div>
  <div style="padding:12px;background:#f0f4f8;border:1px solid #dde;
              border-top:none;border-radius:0 0 8px 8px;
              font-size:11px;color:#888;text-align:center">
    Sent by Retail Intelligence MCP Server
  </div>
</body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = gmail_user
    msg["To"]      = to_email
    msg.attach(MIMEText(body_text, "plain"))
    msg.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(gmail_user, gmail_pass)
            server.sendmail(gmail_user, to_email, msg.as_string())
        return f"Email sent to {to_email} — subject: {subject}"
    except smtplib.SMTPAuthenticationError:
        return "Authentication failed. Check your App Password in the .env file."
    except Exception as e:
        return f"Email failed: {str(e)}"

# ══ RESOURCE: Full catalogue ══════════════════════════════
@mcp.resource("catalogue://current")
def full_catalogue() -> str:
    """Full product catalogue — use this as context when
    answering questions about the product range."""
    results = query("SELECT * FROM products")
    return json.dumps(results, indent=2)

# ══ PROMPT: Viral trend response workflow ════════════════
# YOUR PROMPT — excellent orchestration. Kept as-is.
@mcp.prompt()
def viral_trend_response_workflow(trend_name: str, description: str) -> str:
    """A workflow that handles unexpected fashion trend spikes
    from data capture through to business action execution."""
    return (
        f"You are an aggressive, hyper-responsive fashion merchandising planner.\n"
        f"A trend event has been flagged: '{trend_name}' — {description}\n\n"
        f"Execute this sequential workflow:\n"
        f"1. Call capture_and_analyze_trend with trend_name='{trend_name}', "
        f"description='{description}' to log it and check matching stock.\n"
        f"2. Based on the inventory matches returned, synthesise a commercial strategy.\n"
        f"3. Call draft_trend_alert_email to generate a communication draft.\n"
        f"4. Call send_alert_email to dispatch the alert to the buying team.\n"
        f"Be specific. Use the actual matched SKUs and stock numbers in your strategy."
    )

# ══ PROMPT: Weekly trading brief ══════════════════════════
@mcp.prompt()
def weekly_trading_brief(week: str = "W48") -> str:
    """Generate a complete weekly trading brief for the
    buying and merchandising team."""
    return f"""You are a senior fashion merchandise analyst.
Generate a professional weekly trading brief for {week}.

1. WEEKLY HEADLINE — one sentence summarising performance.

2. TOP 3 PERFORMERS
   Use get_weekly_sales. For each: name, units sold, revenue, why it worked.

3. UNDERPERFORMERS
   Use find_underperformers with threshold 45.
   For each: name, sell-through %, stock value at risk, recommendation.

4. STOCK ALERTS
   Use get_low_stock_alerts. Flag any SKU below 3 weeks cover.

5. THREE ACTION POINTS — numbered, specific, actionable.

Be direct. Use actual numbers. No filler sentences."""

# ══ ENTRY POINT ═══════════════════════════════════════════
if __name__ == "__main__":
    mcp.run()
